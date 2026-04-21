from __future__ import annotations
from datetime import datetime, date, timedelta
import os
import json
import subprocess
import sys
import re
import memory as mem


# ─── DEFINIÇÕES DAS FERRAMENTAS PARA O CLAUDE ──────────────────────────────────

TOOL_DEFINITIONS = [
    {
        "name": "save_memory",
        "description": (
            "Guarda uma informação importante na memória persistente do JARVIS. "
            "Usa esta ferramenta quando o utilizador te ensina algo, diz para te lembrares, "
            "ou define um comportamento personalizado. "
            "Categorias: 'facts' (factos gerais), 'preferences' (preferências do utilizador), "
            "'commands' (comandos personalizados/atalhos definidos pelo utilizador)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "category": {
                    "type": "string",
                    "enum": ["facts", "preferences", "commands"],
                    "description": "Categoria da memória a guardar"
                },
                "content": {
                    "type": "string",
                    "description": "O conteúdo a memorizar, em texto claro e descritivo"
                }
            },
            "required": ["category", "content"]
        }
    },
    {
        "name": "read_memory",
        "description": (
            "Lê todas as memórias guardadas. Usa no início de conversas importantes "
            "ou quando precisas de contexto sobre o utilizador, preferências ou comandos definidos."
        ),
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "delete_memory",
        "description": "Remove uma memória específica pelo índice e categoria.",
        "input_schema": {
            "type": "object",
            "properties": {
                "category": {"type": "string", "enum": ["facts", "preferences", "commands"]},
                "index": {"type": "integer", "description": "Índice (0-based) da memória a remover"}
            },
            "required": ["category", "index"]
        }
    },
    {
        "name": "get_datetime",
        "description": "Obtém a data e hora atual.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "list_folder",
        "description": (
            "Lista os ficheiros e pastas num diretório. "
            "Usa quando o utilizador quer saber o que existe numa pasta."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo da pasta"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "read_file",
        "description": (
            "Lê o conteúdo de um ficheiro de texto (Python, JSON, TXT, CSV, etc.). "
            "Para ficheiros Excel usa read_excel. Para PDF usa read_pdf."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo do ficheiro"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "write_file",
        "description": (
            "Escreve ou substitui o conteúdo de um ficheiro de texto. "
            "Usa para editar ficheiros Python, JSON, TXT, CSV, etc. "
            "ATENÇÃO: substitui o conteúdo completo do ficheiro."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo do ficheiro"},
                "content": {"type": "string", "description": "Conteúdo completo a escrever"}
            },
            "required": ["path", "content"]
        }
    },
    {
        "name": "replace_in_file",
        "description": (
            "Substitui uma string específica por outra num ficheiro de texto. "
            "Mais seguro que write_file — só altera o que é indicado."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo do ficheiro"},
                "old_text": {"type": "string", "description": "Texto a encontrar e substituir"},
                "new_text": {"type": "string", "description": "Texto novo para colocar no lugar"}
            },
            "required": ["path", "old_text", "new_text"]
        }
    },
    {
        "name": "read_excel",
        "description": (
            "Lê um ficheiro Excel (.xlsx, .xls) e devolve o conteúdo em formato legível. "
            "Pode especificar a folha (sheet) a ler."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo do ficheiro Excel"},
                "sheet": {"type": "string", "description": "Nome da folha (opcional, usa a primeira por defeito)"},
                "max_rows": {"type": "integer", "description": "Número máximo de linhas a ler (padrão: 50)"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "create_folder",
        "description": "Cria uma pasta (e subpastas) no caminho indicado.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo da pasta a criar"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "run_command",
        "description": (
            "Executa um comando no terminal (cmd/bash). "
            "Usa para: correr scripts Python, instalar pacotes com pip, "
            "testar se o projeto funciona, etc. "
            "Devolve o output (stdout + stderr). Timeout de 30 segundos."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "command": {"type": "string", "description": "Comando a executar, ex: 'python main.py' ou 'pip install requests'"},
                "cwd": {"type": "string", "description": "Pasta onde executar o comando (opcional)"}
            },
            "required": ["command"]
        }
    },
    {
        "name": "obsidian_search",
        "description": (
            "Pesquisa notas no Obsidian vault por palavra-chave. "
            "Devolve os ficheiros que contêm o termo pesquisado."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Texto a pesquisar nas notas"}
            },
            "required": ["query"]
        }
    },
    {
        "name": "obsidian_read",
        "description": "Lê o conteúdo de uma nota do Obsidian pelo nome do ficheiro.",
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {"type": "string", "description": "Nome do ficheiro .md (ex: 'Reunião.md' ou só 'Reunião')"}
            },
            "required": ["filename"]
        }
    },
    {
        "name": "obsidian_create",
        "description": (
            "Cria ou atualiza uma nota no Obsidian. "
            "Usa para guardar resumos, tarefas, notas de projetos, etc."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {"type": "string", "description": "Nome do ficheiro (sem .md, ex: 'Projeto JARVIS')"},
                "content": {"type": "string", "description": "Conteúdo da nota em Markdown"},
                "folder": {"type": "string", "description": "Subpasta dentro do vault (opcional)"}
            },
            "required": ["filename", "content"]
        }
    },
    {
        "name": "obsidian_list",
        "description": "Lista todas as notas do Obsidian vault.",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "edit_excel_cell",
        "description": "Edita o valor de uma célula específica num ficheiro Excel.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo do ficheiro Excel"},
                "sheet": {"type": "string", "description": "Nome da folha (opcional)"},
                "cell": {"type": "string", "description": "Referência da célula, ex: 'A1', 'B3'"},
                "value": {"type": "string", "description": "Novo valor para a célula"}
            },
            "required": ["path", "cell", "value"]
        }
    },
    # ── APPS & SISTEMA ─────────────────────────────────────────────────────────
    {
        "name": "open_app",
        "description": (
            "Abre uma aplicação no Windows. "
            "Nomes suportados: chrome, teams, notepad, word, excel, vscode, explorer, "
            "calculadora, spotify, outlook. "
            "Também aceita um caminho personalizado via 'path'."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Nome da aplicação (ex: 'chrome', 'notepad', 'word')"},
                "path": {"type": "string", "description": "Caminho completo do executável (opcional)"}
            },
            "required": ["name"]
        }
    },
    {
        "name": "send_email",
        "description": (
            "Envia um email via SMTP. "
            "Requer variáveis de ambiente: EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT, "
            "EMAIL_USERNAME, EMAIL_PASSWORD, EMAIL_FROM, EMAIL_DEFAULT_TO."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "to": {"type": "string", "description": "Endereço de email do destinatário"},
                "subject": {"type": "string", "description": "Assunto do email"},
                "body": {"type": "string", "description": "Corpo do email"}
            },
            "required": ["to", "subject", "body"]
        }
    },
    {
        "name": "web_search",
        "description": (
            "Pesquisa na web usando DuckDuckGo (sem API key). "
            "Devolve resumo, resposta direta e tópicos relacionados."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Pesquisa a realizar"}
            },
            "required": ["query"]
        }
    },
    {
        "name": "web_fetch",
        "description": (
            "Acede a um URL e extrai o texto legível da página. "
            "Útil para ler artigos, documentação, páginas web."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {"type": "string", "description": "URL completo a aceder (ex: https://example.com)"}
            },
            "required": ["url"]
        }
    },
    {
        "name": "notify",
        "description": "Mostra uma notificação no ambiente de trabalho do Windows.",
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Título da notificação"},
                "message": {"type": "string", "description": "Mensagem da notificação"}
            },
            "required": ["title", "message"]
        }
    },
    {
        "name": "set_volume",
        "description": (
            "Define o volume do sistema Windows (0-100) ou silencia/ativa o som. "
            "Ações suportadas: 'mute', 'unmute', ou sem ação para definir nível."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "level": {"type": "integer", "description": "Nível de volume de 0 a 100"},
                "action": {"type": "string", "description": "Ação opcional: 'mute' ou 'unmute'"}
            }
        }
    },
    {
        "name": "get_volume",
        "description": "Obtém o nível atual de volume do sistema Windows (0-100%).",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "play_media",
        "description": (
            "Abre um ficheiro multimédia local ou URL com o leitor padrão do sistema. "
            "Suporta vídeo, áudio, imagens e URLs de streaming."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho local do ficheiro ou URL a abrir"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "read_pdf",
        "description": (
            "Extrai texto de um ficheiro PDF. "
            "Por defeito lê as primeiras 5 páginas."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Caminho completo do ficheiro PDF"},
                "pages": {"type": "integer", "description": "Número de páginas a ler (padrão: 5)"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "calendar_add",
        "description": (
            "Adiciona um evento ao calendário local do JARVIS. "
            "O calendário é guardado num ficheiro JSON local."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Título do evento"},
                "date": {"type": "string", "description": "Data no formato YYYY-MM-DD ou DD/MM/YYYY"},
                "time": {"type": "string", "description": "Hora no formato HH:MM (opcional)"},
                "description": {"type": "string", "description": "Descrição adicional do evento (opcional)"}
            },
            "required": ["title", "date"]
        }
    },
    {
        "name": "calendar_list",
        "description": (
            "Lista eventos do calendário local do JARVIS. "
            "Por defeito mostra eventos futuros (hoje em diante)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "filter": {
                    "type": "string",
                    "description": "Filtro: 'all' (todos), 'today' (hoje), 'week' (esta semana). Padrão: upcoming (futuros)"
                }
            }
        }
    },
    {
        "name": "calendar_delete",
        "description": "Remove um evento do calendário local pelo índice ou pelo título.",
        "input_schema": {
            "type": "object",
            "properties": {
                "index": {"type": "integer", "description": "Índice (0-based) do evento a remover"},
                "title": {"type": "string", "description": "Título (ou parte do título) do evento a remover"}
            }
        }
    },
    # ── SCREENSHOT & VISÃO ─────────────────────────────────────────────────────
    {
        "name": "take_screenshot",
        "description": (
            "Tira uma screenshot do ecrã atual e analisa o que está visível. "
            "Usa quando o utilizador quer saber o que está no ecrã, "
            "analisar um erro, ou ver o estado de uma aplicação."
        ),
        "input_schema": {"type": "object", "properties": {}}
    },
    # ── BROWSER (Playwright) ───────────────────────────────────────────────────
    {
        "name": "browser_navigate",
        "description": (
            "Abre o browser e navega para um URL. "
            "Usa para abrir websites, fazer pesquisas, aceder a serviços web."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {"type": "string", "description": "URL a navegar (ex: https://google.com)"}
            },
            "required": ["url"]
        }
    },
    {
        "name": "browser_click",
        "description": (
            "Clica num elemento da página do browser pelo texto visível ou seletor CSS. "
            "Usa após browser_navigate para interagir com a página."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "selector_or_text": {"type": "string", "description": "Texto visível do elemento ou seletor CSS (ex: 'Pesquisar', '#btn-submit')"}
            },
            "required": ["selector_or_text"]
        }
    },
    {
        "name": "browser_fill",
        "description": (
            "Preenche um campo de formulário no browser. "
            "Usa o label, placeholder ou seletor CSS do campo."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "selector_or_label": {"type": "string", "description": "Label, placeholder ou seletor CSS do campo"},
                "value": {"type": "string", "description": "Valor a preencher no campo"}
            },
            "required": ["selector_or_label", "value"]
        }
    },
    {
        "name": "browser_screenshot",
        "description": (
            "Tira uma screenshot da página atual do browser e analisa o que está visível. "
            "Útil para verificar o estado da página após navegar ou interagir."
        ),
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "browser_get_text",
        "description": "Obtém o texto visível da página atual do browser.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "browser_close",
        "description": "Fecha o browser Playwright.",
        "input_schema": {"type": "object", "properties": {}}
    },
    # ── TAREFAS AGENDADAS ──────────────────────────────────────────────────────
    {
        "name": "schedule_task",
        "description": (
            "Cria uma nova tarefa agendada para o JARVIS executar automaticamente. "
            "O horário pode ser em linguagem natural ('todos os dias às 8h', 'de hora a hora', "
            "'às segundas às 9h') ou em formato cron ('0 8 * * *'). "
            "O comando é o que o JARVIS vai fazer quando a tarefa disparar."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Nome descritivo da tarefa (ex: 'Relatório diário')"},
                "schedule": {"type": "string", "description": "Horário em linguagem natural ou cron (ex: 'todos os dias às 8h', '0 8 * * *')"},
                "command": {"type": "string", "description": "Instrução para o JARVIS executar (ex: 'Verifica o email e dá-me um resumo')"}
            },
            "required": ["name", "schedule", "command"]
        }
    },
    {
        "name": "list_scheduled_tasks",
        "description": "Lista todas as tarefas agendadas do JARVIS.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "delete_scheduled_task",
        "description": "Remove uma tarefa agendada pelo ID ou nome.",
        "input_schema": {
            "type": "object",
            "properties": {
                "id_or_name": {"type": "string", "description": "ID ou nome da tarefa a remover"}
            },
            "required": ["id_or_name"]
        }
    },
    {
        "name": "toggle_scheduled_task",
        "description": "Ativa ou pausa uma tarefa agendada pelo ID ou nome.",
        "input_schema": {
            "type": "object",
            "properties": {
                "id_or_name": {"type": "string", "description": "ID ou nome da tarefa a ativar/pausar"}
            },
            "required": ["id_or_name"]
        }
    },
]


# ─── EXECUÇÃO DAS FERRAMENTAS ──────────────────────────────────────────────────

def execute_tool(name: str, tool_input: dict) -> str:
    if name == "save_memory":
        return mem.save_memory(tool_input["category"], tool_input["content"])

    if name == "read_memory":
        return mem.read_memory()

    if name == "delete_memory":
        return mem.delete_memory(tool_input["category"], tool_input["index"])

    if name == "get_datetime":
        now = datetime.now()
        return f"Data: {now.strftime('%d/%m/%Y')}, Hora: {now.strftime('%H:%M:%S')}, Dia: {now.strftime('%A')}"

    if name == "create_folder":
        return _create_folder(tool_input["path"])

    if name == "run_command":
        return _run_command(tool_input["command"], tool_input.get("cwd"))

    if name == "list_folder":
        return _list_folder(tool_input["path"])

    if name == "read_file":
        return _read_file(tool_input["path"])

    if name == "write_file":
        return _write_file(tool_input["path"], tool_input["content"])

    if name == "replace_in_file":
        return _replace_in_file(tool_input["path"], tool_input["old_text"], tool_input["new_text"])

    if name == "read_excel":
        return _read_excel(
            tool_input["path"],
            tool_input.get("sheet"),
            tool_input.get("max_rows", 50)
        )

    if name == "obsidian_search":
        return _obsidian_search(tool_input["query"])

    if name == "obsidian_read":
        return _obsidian_read(tool_input["filename"])

    if name == "obsidian_create":
        return _obsidian_create(tool_input["filename"], tool_input["content"], tool_input.get("folder", ""))

    if name == "obsidian_list":
        return _obsidian_list()

    if name == "edit_excel_cell":
        return _edit_excel_cell(
            tool_input["path"],
            tool_input.get("sheet"),
            tool_input["cell"],
            tool_input["value"]
        )

    if name == "open_app":
        return _open_app(tool_input["name"], tool_input.get("path"))

    if name == "send_email":
        return _send_email(tool_input["to"], tool_input["subject"], tool_input["body"])

    if name == "web_search":
        return _web_search(tool_input["query"])

    if name == "web_fetch":
        return _web_fetch(tool_input["url"])

    if name == "notify":
        return _notify(tool_input["title"], tool_input["message"])

    if name == "set_volume":
        return _set_volume(tool_input.get("level"), tool_input.get("action"))

    if name == "get_volume":
        return _get_volume()

    if name == "play_media":
        return _play_media(tool_input["path"])

    if name == "read_pdf":
        return _read_pdf(tool_input["path"], tool_input.get("pages", 5))

    if name == "calendar_add":
        return _calendar_add(
            tool_input["title"],
            tool_input["date"],
            tool_input.get("time", ""),
            tool_input.get("description", "")
        )

    if name == "calendar_list":
        return _calendar_list(tool_input.get("filter", "upcoming"))

    if name == "calendar_delete":
        return _calendar_delete(tool_input.get("index"), tool_input.get("title"))

    # ── Screenshot & Computer Vision ───────────────────────────────────────────
    if name == "take_screenshot":
        return _take_screenshot()

    # ── Browser tools ──────────────────────────────────────────────────────────
    if name == "browser_navigate":
        try:
            from browser_tools import browser_navigate
            return browser_navigate(tool_input["url"])
        except ImportError:
            return "browser_tools não disponível."

    if name == "browser_click":
        try:
            from browser_tools import browser_click
            return browser_click(tool_input["selector_or_text"])
        except ImportError:
            return "browser_tools não disponível."

    if name == "browser_fill":
        try:
            from browser_tools import browser_fill
            return browser_fill(tool_input["selector_or_label"], tool_input["value"])
        except ImportError:
            return "browser_tools não disponível."

    if name == "browser_screenshot":
        try:
            from browser_tools import browser_screenshot
            return browser_screenshot()
        except ImportError:
            return "browser_tools não disponível."

    if name == "browser_get_text":
        try:
            from browser_tools import browser_get_text
            return browser_get_text()
        except ImportError:
            return "browser_tools não disponível."

    if name == "browser_close":
        try:
            from browser_tools import browser_close
            return browser_close()
        except ImportError:
            return "browser_tools não disponível."

    # ── Scheduled tasks ────────────────────────────────────────────────────────
    if name == "schedule_task":
        try:
            from scheduler import jarvis_scheduler
            return jarvis_scheduler.create_task(
                tool_input["name"],
                tool_input["schedule"],
                tool_input["command"]
            )
        except ImportError:
            return "APScheduler não instalado. Corre: pip install apscheduler"

    if name == "list_scheduled_tasks":
        try:
            from scheduler import jarvis_scheduler
            return jarvis_scheduler.list_tasks()
        except ImportError:
            return "APScheduler não instalado. Corre: pip install apscheduler"

    if name == "delete_scheduled_task":
        try:
            from scheduler import jarvis_scheduler
            return jarvis_scheduler.delete_task(tool_input["id_or_name"])
        except ImportError:
            return "APScheduler não instalado. Corre: pip install apscheduler"

    if name == "toggle_scheduled_task":
        try:
            from scheduler import jarvis_scheduler
            return jarvis_scheduler.toggle_task(tool_input["id_or_name"])
        except ImportError:
            return "APScheduler não instalado. Corre: pip install apscheduler"

    return f"Ferramenta '{name}' não reconhecida."


# ─── IMPLEMENTAÇÕES EXISTENTES ─────────────────────────────────────────────────

def _create_folder(path: str) -> str:
    path = os.path.expandvars(path)
    try:
        os.makedirs(path, exist_ok=True)
        return f"Pasta criada: {path}"
    except Exception as e:
        return f"Erro ao criar pasta: {e}"


def _run_command(command: str, cwd: str | None) -> str:
    if cwd:
        cwd = os.path.expandvars(cwd)
    try:
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            timeout=30,
            cwd=cwd or None
        )
        output = ""
        if result.stdout:
            output += f"OUTPUT:\n{result.stdout}"
        if result.stderr:
            output += f"\nERROS:\n{result.stderr}"
        if not output:
            output = "(sem output)"
        output += f"\nCódigo de saída: {result.returncode}"
        return output[:3000]
    except subprocess.TimeoutExpired:
        return "Timeout: comando demorou mais de 30 segundos."
    except Exception as e:
        return f"Erro ao executar comando: {e}"


def _list_folder(path: str) -> str:
    path = os.path.expandvars(path)
    if not os.path.exists(path):
        return f"Pasta não encontrada: {path}"
    items = []
    try:
        for entry in sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower())):
            prefix = "📁" if entry.is_dir() else "📄"
            size = ""
            if entry.is_file():
                s = entry.stat().st_size
                size = f" ({s/1024:.1f} KB)" if s > 1024 else f" ({s} B)"
            items.append(f"{prefix} {entry.name}{size}")
    except PermissionError:
        return f"Sem permissão para aceder a: {path}"
    return f"Conteúdo de '{path}':\n" + "\n".join(items) if items else "Pasta vazia."


def _read_file(path: str) -> str:
    path = os.path.expandvars(path)
    if not os.path.exists(path):
        return f"Ficheiro não encontrado: {path}"
    try:
        size = os.path.getsize(path)
        if size > 100_000:
            return f"Ficheiro demasiado grande ({size/1024:.0f} KB). Máximo: 100 KB."
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        return f"=== {path} ===\n{content}"
    except Exception as e:
        return f"Erro ao ler ficheiro: {e}"


def _write_file(path: str, content: str) -> str:
    path = os.path.expandvars(path)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"Ficheiro guardado com sucesso: {path}"
    except Exception as e:
        return f"Erro ao guardar ficheiro: {e}"


def _replace_in_file(path: str, old_text: str, new_text: str) -> str:
    path = os.path.expandvars(path)
    if not os.path.exists(path):
        return f"Ficheiro não encontrado: {path}"
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        if old_text not in content:
            return f"Texto não encontrado no ficheiro: '{old_text}'"
        new_content = content.replace(old_text, new_text)
        with open(path, "w", encoding="utf-8") as f:
            f.write(new_content)
        count = content.count(old_text)
        return f"Substituição feita com sucesso ({count} ocorrência(s)) em: {path}"
    except Exception as e:
        return f"Erro: {e}"


def _read_excel(path: str, sheet: str | None, max_rows: int) -> str:
    path = os.path.expandvars(path)
    try:
        import openpyxl
    except ImportError:
        return "openpyxl não instalado. Corre: pip install openpyxl"
    if not os.path.exists(path):
        return f"Ficheiro não encontrado: {path}"
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        sheets_info = f"Folhas disponíveis: {', '.join(wb.sheetnames)}\n"
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append(f"... (limitado a {max_rows} linhas)")
                break
            rows.append("\t".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return sheets_info + f"Folha: '{ws.title}'\n" + "\n".join(rows)
    except Exception as e:
        return f"Erro ao ler Excel: {e}"


OBSIDIAN_VAULT = r"C:\Users\vitor.leonardo\Documents\Obsidian Vault"


def _obsidian_list() -> str:
    notes = []
    for root, _, files in os.walk(OBSIDIAN_VAULT):
        for f in files:
            if f.endswith(".md"):
                rel = os.path.relpath(os.path.join(root, f), OBSIDIAN_VAULT)
                notes.append(rel)
    if not notes:
        return "Vault vazio."
    return f"Notas no Obsidian ({len(notes)}):\n" + "\n".join(sorted(notes))


def _obsidian_search(query: str) -> str:
    results = []
    q = query.lower()
    for root, _, files in os.walk(OBSIDIAN_VAULT):
        for f in files:
            if not f.endswith(".md"):
                continue
            path = os.path.join(root, f)
            try:
                with open(path, "r", encoding="utf-8") as fp:
                    content = fp.read()
                if q in content.lower() or q in f.lower():
                    rel = os.path.relpath(path, OBSIDIAN_VAULT)
                    idx = content.lower().find(q)
                    snippet = content[max(0,idx-60):idx+100].replace("\n"," ").strip()
                    results.append(f"📄 {rel}\n   ...{snippet}...")
            except Exception:
                pass
    if not results:
        return f"Nenhuma nota encontrada com '{query}'."
    return f"Resultados para '{query}':\n\n" + "\n\n".join(results[:10])


def _obsidian_read(filename: str) -> str:
    if not filename.endswith(".md"):
        filename += ".md"
    for root, _, files in os.walk(OBSIDIAN_VAULT):
        for f in files:
            if f.lower() == filename.lower():
                path = os.path.join(root, f)
                with open(path, "r", encoding="utf-8") as fp:
                    return f"=== {f} ===\n{fp.read()}"
    return f"Nota '{filename}' não encontrada no vault."


def _obsidian_create(filename: str, content: str, folder: str) -> str:
    if not filename.endswith(".md"):
        filename += ".md"
    if folder:
        dest_dir = os.path.join(OBSIDIAN_VAULT, folder)
    else:
        dest_dir = OBSIDIAN_VAULT
    os.makedirs(dest_dir, exist_ok=True)
    path = os.path.join(dest_dir, filename)
    with open(path, "w", encoding="utf-8") as fp:
        fp.write(content)
    return f"Nota criada/atualizada: {os.path.relpath(path, OBSIDIAN_VAULT)}"


def _edit_excel_cell(path: str, sheet: str | None, cell: str, value: str) -> str:
    path = os.path.expandvars(path)
    try:
        import openpyxl
    except ImportError:
        return "openpyxl não instalado. Corre: pip install openpyxl"
    if not os.path.exists(path):
        return f"Ficheiro não encontrado: {path}"
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet] if sheet else wb.active
        old_value = ws[cell].value
        try:
            ws[cell] = float(value) if "." in value else int(value)
        except ValueError:
            ws[cell] = value
        wb.save(path)
        wb.close()
        return f"Célula {cell} em '{ws.title}' alterada: '{old_value}' → '{value}' em {path}"
    except Exception as e:
        return f"Erro ao editar Excel: {e}"


# ─── NOVAS IMPLEMENTAÇÕES ──────────────────────────────────────────────────────

_APP_MAP = {
    "chrome":      ("exe",  "chrome.exe"),
    "teams":       ("url",  "ms-teams://"),
    "notepad":     ("exe",  "notepad.exe"),
    "word":        ("exe",  "winword.exe"),
    "excel":       ("exe",  "excel.exe"),
    "vscode":      ("exe",  "code"),
    "explorer":    ("exe",  "explorer.exe"),
    "calculadora": ("exe",  "calc.exe"),
    "calc":        ("exe",  "calc.exe"),
    "spotify":     ("exe",  "spotify.exe"),
    "outlook":     ("exe",  "outlook.exe"),
}

_APP_FALLBACK_PATHS = {
    "outlook": [
        r"C:\Program Files\Microsoft Office\root\Office16\OUTLOOK.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\OUTLOOK.EXE",
        r"C:\Program Files\Microsoft Office\Office16\OUTLOOK.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office16\OUTLOOK.EXE",
        r"C:\Program Files\Microsoft Office\Office15\OUTLOOK.EXE",
    ],
    "word": [
        r"C:\Program Files\Microsoft Office\root\Office16\WINWORD.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\WINWORD.EXE",
    ],
    "excel": [
        r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE",
    ],
    "teams": [
        r"C:\Users\vitor.leonardo\AppData\Local\Microsoft\Teams\current\Teams.exe",
        r"C:\Program Files\Microsoft\Teams\current\Teams.exe",
    ],
    "spotify": [
        r"C:\Users\vitor.leonardo\AppData\Roaming\Spotify\Spotify.exe",
    ],
}


def _open_app(name: str, path: str | None) -> str:
    try:
        if path:
            path = os.path.expandvars(path)
            subprocess.Popen(path)
            return f"Aplicação aberta: {path}"

        key = name.lower().strip()
        if key not in _APP_MAP:
            try:
                subprocess.Popen(name)
                return f"Aplicação '{name}' iniciada."
            except Exception:
                available = ", ".join(sorted(_APP_MAP.keys()))
                return (
                    f"Aplicação '{name}' não reconhecida.\n"
                    f"Nomes suportados: {available}\n"
                    f"Ou forneça o parâmetro 'path' com o caminho completo."
                )

        kind, target = _APP_MAP[key]
        if kind == "url":
            subprocess.run(f"start {target}", shell=True)
            return f"A abrir {name}..."

        fallbacks = _APP_FALLBACK_PATHS.get(key, [])
        for fb_path in fallbacks:
            if os.path.exists(fb_path):
                subprocess.Popen([fb_path])
                return f"A abrir {name}..."

        check = subprocess.run(f"where {target}", shell=True,
                               capture_output=True, text=True)
        if check.returncode == 0:
            subprocess.Popen(target, shell=True)
            return f"A abrir {name}..."

        return (f"Não foi possível encontrar '{name}'. "
                f"Indica o caminho completo do executável.")

    except Exception as e:
        return f"Erro ao abrir '{name}': {e}"


def _send_email(to: str, subject: str, body: str) -> str:
    import urllib.parse

    default_to = os.environ.get("EMAIL_DEFAULT_TO", "")
    recipient  = to if to else default_to

    try:
        import win32com.client
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To      = recipient
        mail.Subject = subject
        mail.Body    = body
        mail.Send()
        return f"Email enviado para {recipient} via Outlook."
    except ImportError:
        pass
    except Exception:
        pass

    try:
        params = urllib.parse.urlencode({
            "subject": subject,
            "body":    body
        }, quote_via=urllib.parse.quote)
        mailto_url = f"mailto:{urllib.parse.quote(recipient)}?{params}"
        subprocess.run(f'start "" "{mailto_url}"', shell=True)
        return (f"Email preparado e aberto no teu cliente de email.\n"
                f"Para: {recipient} | Assunto: {subject}\n"
                f"Clica em Enviar para confirmar.")
    except Exception:
        pass

    smtp_server  = os.environ.get("EMAIL_SMTP_SERVER", "")
    username     = os.environ.get("EMAIL_USERNAME", "")
    password     = os.environ.get("EMAIL_PASSWORD", "")
    from_addr    = os.environ.get("EMAIL_FROM", username)

    if not smtp_server or not username or not password:
        return (
            "Não foi possível enviar o email. Opções:\n"
            "1. Instala pywin32: pip install pywin32  (usa o Outlook instalado)\n"
            "2. Configura SMTP no ficheiro .env com EMAIL_SMTP_SERVER, EMAIL_USERNAME, EMAIL_PASSWORD"
        )

    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        smtp_port = int(os.environ.get("EMAIL_SMTP_PORT", "587"))
        msg = MIMEMultipart()
        msg["From"] = from_addr
        msg["To"] = recipient
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))

        with smtplib.SMTP(smtp_server, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(username, password)
            server.sendmail(from_addr, recipient, msg.as_string())

        return f"Email enviado com sucesso para {recipient}.\nAssunto: {subject}"
    except Exception as e:
        return f"Erro ao enviar email: {e}"


def _web_search(query: str) -> str:
    try:
        import requests as req
    except ImportError:
        return "requests não instalado. Corre: pip install requests"

    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

    try:
        import urllib.parse
        encoded = urllib.parse.quote_plus(query)
        url = f"https://api.duckduckgo.com/?q={encoded}&format=json&no_html=1&skip_disambig=1"
        resp = req.get(url, headers=headers, timeout=8)
        data = resp.json()

        parts = []
        abstract = data.get("AbstractText", "").strip()
        answer = data.get("Answer", "").strip()
        source = data.get("AbstractSource", "").strip()
        source_url = data.get("AbstractURL", "").strip()

        if answer:
            parts.append(f"Resposta direta: {answer}")
        if abstract:
            src_info = f" (Fonte: {source} — {source_url})" if source else ""
            parts.append(f"Resumo: {abstract}{src_info}")

        related = data.get("RelatedTopics", [])[:5]
        if related:
            rel_lines = []
            for t in related:
                if isinstance(t, dict) and t.get("Text"):
                    text = t["Text"][:120]
                    link = t.get("FirstURL", "")
                    rel_lines.append(f"  - {text}" + (f"\n    {link}" if link else ""))
            if rel_lines:
                parts.append("Tópicos relacionados:\n" + "\n".join(rel_lines))

        if parts:
            return f"Resultados para '{query}':\n\n" + "\n\n".join(parts)

    except Exception:
        pass

    try:
        import urllib.parse
        encoded = urllib.parse.quote_plus(query)
        url = f"https://html.duckduckgo.com/html/?q={encoded}"
        resp = req.get(url, headers=headers, timeout=8)
        html = resp.text

        results = []
        search_tag = 'class="result__a"'
        pos = 0
        while len(results) < 5:
            idx = html.find(search_tag, pos)
            if idx == -1:
                break
            href_start = html.rfind("href=", max(0, idx - 200), idx)
            if href_start != -1:
                q_char = html[href_start + 5]
                href_end = html.find(q_char, href_start + 6)
                href = html[href_start + 6:href_end] if href_end != -1 else ""
            else:
                href = ""
            tag_end = html.find(">", idx)
            close_a = html.find("</a>", tag_end)
            link_text = re.sub(r"<[^>]+>", "", html[tag_end + 1:close_a]).strip() if tag_end != -1 and close_a != -1 else ""
            if link_text:
                results.append(f"  - {link_text}\n    {href}" if href else f"  - {link_text}")
            pos = idx + len(search_tag)

        if results:
            return f"Resultados para '{query}' (DuckDuckGo):\n\n" + "\n\n".join(results)
        return f"Nenhum resultado encontrado para '{query}'."

    except Exception as e:
        return f"Erro na pesquisa web: {e}"


def _web_fetch(url: str) -> str:
    try:
        import requests as req
    except ImportError:
        return "requests não instalado. Corre: pip install requests"

    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        resp = req.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        html = resp.text
        text = re.sub(r"<[^>]+>", "", html)
        text = re.sub(r"\n\s*\n\s*\n", "\n\n", text)
        text = re.sub(r"[ \t]{2,}", " ", text)
        text = text.strip()
        if len(text) > 3000:
            text = text[:3000] + "\n\n[... conteúdo truncado a 3000 caracteres]"
        return f"Conteúdo de {url}:\n\n{text}"
    except Exception as e:
        return f"Erro ao aceder a '{url}': {e}"


def _notify(title: str, message: str) -> str:
    try:
        from plyer import notification
        notification.notify(
            title=title,
            message=message,
            app_name="JARVIS",
            timeout=5
        )
        return f"Notificação enviada: '{title}'"
    except ImportError:
        pass
    except Exception:
        pass

    try:
        ps_script = (
            f"Add-Type -AssemblyName System.Windows.Forms; "
            f"$notify = New-Object System.Windows.Forms.NotifyIcon; "
            f"$notify.Icon = [System.Drawing.SystemIcons]::Information; "
            f"$notify.Visible = $true; "
            f"$notify.ShowBalloonTip(5000, '{title}', '{message}', "
            f"[System.Windows.Forms.ToolTipIcon]::Info)"
        )
        subprocess.Popen(
            ["powershell", "-WindowStyle", "Hidden", "-Command", ps_script],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        return f"Notificação enviada (PowerShell): '{title}'"
    except Exception:
        pass

    print(f"\n[JARVIS NOTIFICAÇÃO] {title}: {message}\n")
    return f"Notificação impressa no terminal: '{title}' — {message}"


def _set_volume(level: int | None, action: str | None) -> str:
    try:
        from ctypes import cast, POINTER
        from comtypes import CLSCTX_ALL
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume

        devices = AudioUtilities.GetSpeakers()
        interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
        vol = cast(interface, POINTER(IAudioEndpointVolume))

        if action == "mute":
            vol.SetMute(1, None)
            return "Volume silenciado (mute)."
        elif action == "unmute":
            vol.SetMute(0, None)
            return "Volume ativado (unmute)."
        elif level is not None:
            clamped = max(0, min(100, level))
            vol.SetMasterVolumeLevelScalar(clamped / 100.0, None)
            return f"Volume definido para {clamped}%."
        else:
            return "Especifica 'level' (0-100) ou 'action' (mute/unmute)."

    except ImportError:
        pass
    except Exception:
        pass

    try:
        if action == "mute":
            subprocess.run(
                'powershell -c "$wsh=New-Object -comObject WScript.Shell; $wsh.SendKeys([char]173)"',
                shell=True, timeout=5
            )
            return "Mute ativado (via PowerShell)."
        elif action == "unmute":
            subprocess.run(
                'powershell -c "$wsh=New-Object -comObject WScript.Shell; $wsh.SendKeys([char]173)"',
                shell=True, timeout=5
            )
            return "Unmute ativado (via PowerShell — toggle mute)."
        elif level is not None:
            clamped = max(0, min(100, level))
            result = subprocess.run(
                f'powershell -c "(New-Object -com Shell.Application).Windows() | '
                f'ForEach-Object {{}}; '
                f'$vol = {clamped / 100.0}; '
                f'$wsh = New-Object -comObject WScript.Shell; '
                f'1..50 | % {{$wsh.SendKeys([char]174)}}; '
                f'[int]($vol * 50) | % {{ 1..$_ }} | % {{$wsh.SendKeys([char]175)}}"',
                shell=True, timeout=10
            )
            return f"Volume aproximado para {clamped}% (via PowerShell — instala pycaw para controlo preciso: pip install pycaw comtypes)."
        else:
            return "Especifica 'level' (0-100) ou 'action' (mute/unmute). Instala pycaw para melhor suporte: pip install pycaw comtypes"
    except Exception as e:
        return f"Erro ao definir volume: {e}\nInstala pycaw: pip install pycaw comtypes"


def _get_volume() -> str:
    try:
        from ctypes import cast, POINTER
        from comtypes import CLSCTX_ALL
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume

        devices = AudioUtilities.GetSpeakers()
        interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
        vol = cast(interface, POINTER(IAudioEndpointVolume))
        current = vol.GetMasterVolumeLevelScalar() * 100
        muted = vol.GetMute()
        mute_str = " (silenciado)" if muted else ""
        return f"Volume atual: {current:.0f}%{mute_str}"
    except ImportError:
        return "pycaw não instalado. Instala com: pip install pycaw comtypes"
    except Exception as e:
        return f"Erro ao obter volume: {e}"


def _play_media(path: str) -> str:
    path_expanded = os.path.expandvars(path)
    try:
        if path_expanded.startswith("http://") or path_expanded.startswith("https://") or "://" in path_expanded:
            subprocess.run(f'start "" "{path_expanded}"', shell=True)
            return f"A abrir URL no leitor padrão: {path_expanded}"
        if os.path.exists(path_expanded):
            os.startfile(path_expanded)
            return f"A reproduzir ficheiro: {path_expanded}"
        os.startfile(path_expanded)
        return f"A abrir: {path_expanded}"
    except Exception as e:
        return f"Erro ao abrir média '{path}': {e}"


def _read_pdf(path: str, pages: int) -> str:
    path = os.path.expandvars(path)
    if not os.path.exists(path):
        return f"Ficheiro não encontrado: {path}"

    try:
        import pdfplumber
        texts = []
        with pdfplumber.open(path) as pdf:
            total = len(pdf.pages)
            limit = min(pages, total)
            for i in range(limit):
                page_text = pdf.pages[i].extract_text()
                if page_text:
                    texts.append(f"--- Página {i+1} ---\n{page_text}")
        result = "\n\n".join(texts)
        if len(result) > 5000:
            result = result[:5000] + "\n\n[... truncado a 5000 caracteres]"
        header = f"PDF: {os.path.basename(path)} ({total} páginas total, lidas {limit})\n\n"
        return header + (result if result else "(Sem texto extraível nas páginas lidas)")
    except ImportError:
        pass
    except Exception as e:
        return f"Erro ao ler PDF com pdfplumber: {e}"

    try:
        import PyPDF2
        texts = []
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            total = len(reader.pages)
            limit = min(pages, total)
            for i in range(limit):
                page_text = reader.pages[i].extract_text()
                if page_text:
                    texts.append(f"--- Página {i+1} ---\n{page_text}")
        result = "\n\n".join(texts)
        if len(result) > 5000:
            result = result[:5000] + "\n\n[... truncado a 5000 caracteres]"
        header = f"PDF: {os.path.basename(path)} ({total} páginas total, lidas {limit})\n\n"
        return header + (result if result else "(Sem texto extraível)")
    except ImportError:
        return (
            "Nenhuma biblioteca PDF instalada.\n"
            "Instala pdfplumber: pip install pdfplumber\n"
            "Ou PyPDF2: pip install PyPDF2"
        )
    except Exception as e:
        return f"Erro ao ler PDF: {e}"


# ─── SCREENSHOT ────────────────────────────────────────────────────────────────

def _take_screenshot() -> str:
    """Tira screenshot do monitor principal e devolve JSON com base64."""
    try:
        import mss
        import base64
        from PIL import Image
        import io

        with mss.mss() as sct:
            monitor = sct.monitors[1]  # monitor primário
            screenshot = sct.grab(monitor)
            img = Image.frombytes('RGB', screenshot.size, screenshot.bgra, 'raw', 'BGRX')

            # Redimensiona para máximo 1920x1080 para poupar tokens
            img.thumbnail((1920, 1080), Image.LANCZOS)

            buffer = io.BytesIO()
            img.save(buffer, format='PNG', optimize=True)
            b64 = base64.b64encode(buffer.getvalue()).decode()

            return json.dumps({
                "type": "screenshot",
                "base64": b64,
                "media_type": "image/png",
                "width": img.width,
                "height": img.height
            })
    except ImportError as e:
        missing = str(e).split("'")[1] if "'" in str(e) else str(e)
        return f"Dependência em falta: {missing}. Instala com: pip install mss Pillow"
    except Exception as e:
        return f"Erro ao tirar screenshot: {e}"


# ─── CALENDÁRIO LOCAL ──────────────────────────────────────────────────────────

_CALENDAR_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jarvis_calendar.json")


def _load_calendar() -> dict:
    if os.path.exists(_CALENDAR_FILE):
        try:
            with open(_CALENDAR_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"events": []}


def _save_calendar(data: dict) -> None:
    with open(_CALENDAR_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _parse_date(date_str: str) -> str:
    """Normaliza datas para YYYY-MM-DD."""
    date_str = date_str.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_str


def _calendar_add(title: str, date_str: str, time_str: str, description: str) -> str:
    try:
        data = _load_calendar()
        normalized_date = _parse_date(date_str)
        event = {
            "title": title,
            "date": normalized_date,
            "time": time_str,
            "description": description,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        data["events"].append(event)
        data["events"].sort(key=lambda e: e.get("date", ""))
        _save_calendar(data)
        time_info = f" às {time_str}" if time_str else ""
        return f"Evento adicionado ao calendário:\n'{title}' — {normalized_date}{time_info}"
    except Exception as e:
        return f"Erro ao adicionar evento: {e}"


def _calendar_list(filter_mode: str) -> str:
    try:
        data = _load_calendar()
        events = data.get("events", [])

        if not events:
            return "Calendário vazio. Usa calendar_add para adicionar eventos."

        today_str = date.today().strftime("%Y-%m-%d")
        week_end = (date.today() + timedelta(days=7)).strftime("%Y-%m-%d")

        if filter_mode == "all":
            filtered = events
        elif filter_mode == "today":
            filtered = [e for e in events if e.get("date", "") == today_str]
        elif filter_mode == "week":
            filtered = [e for e in events if today_str <= e.get("date", "") <= week_end]
        else:
            filtered = [e for e in events if e.get("date", "") >= today_str]

        if not filtered:
            labels = {"all": "nenhum evento", "today": "nenhum evento hoje",
                      "week": "nenhum evento esta semana", "upcoming": "nenhum evento futuro"}
            return f"Calendário: {labels.get(filter_mode, 'nenhum evento encontrado')}."

        lines = [f"Calendário ({filter_mode}) — {len(filtered)} evento(s):\n"]
        for i, ev in enumerate(filtered):
            time_info = f" {ev['time']}" if ev.get("time") else ""
            desc = f"\n   {ev['description']}" if ev.get("description") else ""
            global_idx = events.index(ev)
            lines.append(f"[{global_idx}] {ev['date']}{time_info} — {ev['title']}{desc}")

        return "\n".join(lines)
    except Exception as e:
        return f"Erro ao listar eventos: {e}"


def _calendar_delete(index: int | None, title: str | None) -> str:
    try:
        data = _load_calendar()
        events = data.get("events", [])

        if not events:
            return "Calendário vazio."

        if index is not None:
            if index < 0 or index >= len(events):
                return f"Índice {index} inválido. O calendário tem {len(events)} evento(s) (0-{len(events)-1})."
            removed = events.pop(index)
            _save_calendar(data)
            return f"Evento removido: '{removed['title']}' ({removed['date']})"

        if title:
            title_lower = title.lower()
            matches = [(i, e) for i, e in enumerate(events) if title_lower in e.get("title", "").lower()]
            if not matches:
                return f"Nenhum evento encontrado com título contendo '{title}'."
            if len(matches) == 1:
                idx, ev = matches[0]
                events.pop(idx)
                _save_calendar(data)
                return f"Evento removido: '{ev['title']}' ({ev['date']})"
            options = "\n".join(f"  [{i}] {e['date']} — {e['title']}" for i, e in matches)
            return f"Vários eventos encontrados com '{title}':\n{options}\nUsa calendar_delete com 'index' para especificar."

        return "Especifica 'index' ou 'title' para remover um evento."
    except Exception as e:
        return f"Erro ao remover evento: {e}"
